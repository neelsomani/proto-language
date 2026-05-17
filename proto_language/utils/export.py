"""Five tables, each with a single natural format:.

- sequences:    One row per (result_idx, construct, segment)
- constraints:  One row per (result_idx, construct, segment, constraint)
- constructs:   One row per (result_idx, construct)
- optimization: One row per (timepoint, result_idx)
- fasta:        Standard FASTA format for bioinformatics pipelines

Supports CSV, TSV, JSON, FASTA, and Excel output formats.
"""

import copy
import csv
import json
from io import StringIO
from pathlib import Path
from typing import IO, Any, Literal

import numpy as np

from proto_language.utils.helpers import make_json_safe

# Type aliases
Format = Literal["csv", "tsv", "json", "xlsx"]
Results = dict[str, Any]  # Output from build_results()


# =============================================================================
# Build results
# =============================================================================


def build_results(
    constructs: list[Any],
    energy_scores: list[float],
) -> Results:
    """Build standardized results from live Construct objects.

    Produces the canonical format consumed by all flatten/export functions.
    Infinite/NaN energy scores are converted to None for JSON compatibility.
    When set, ``seq.structure`` / ``seq.logits`` are carried as opaque
    ``_structure`` / ``_logits`` entries that external materializers consume.

    Args:
        constructs (list[Any]): List of Construct objects.
        energy_scores (list[float]): List of energy scores (one per result).

    Returns:
        Results: Dict with "results" (list of result dicts) and "best_result_idx"::

            {
                "results": [
                    {
                        "result_idx": 0,
                        "energy_score": 0.5,
                        "constructs": [
                            {
                                "label": "construct_0",
                                "type": "dna",
                                "segments": [
                                    {
                                        "label": "promoter",
                                        "sequence": "ATCG",
                                        "constraints": {
                                            "gc_content": {
                                                "score": 0.5,
                                                "weight": 1.0,
                                                "weighted_score": 0.5,
                                                "data": {"gc_content": 50.0},
                                            }
                                        },
                                        "generators": {"proteinmpnn": {"perplexity": 1.8}},
                                        "metadata": {},
                                    }
                                ],
                            }
                        ],
                    }
                ],
                "best_result_idx": 0,
            }
    """
    if not constructs or not constructs[0].segments:
        return {"results": [], "best_result_idx": 0}

    num_results = len(constructs[0].segments[0].result_sequences)
    results = []

    for result_idx in range(num_results):
        structured_constructs = []
        for construct in constructs:
            structured_segments = []
            for seg_idx, segment in enumerate(construct.segments):
                seq = segment.result_sequences[result_idx]
                seg_dict: dict[str, Any] = {
                    "label": segment.label or f"segment_{seg_idx}",
                    "sequence": seq.sequence,
                    "constraints": make_json_safe(copy.deepcopy(seq._constraints_metadata)),
                    "generators": make_json_safe(copy.deepcopy(seq._generator_metadata)),
                    "metadata": make_json_safe(copy.deepcopy(seq._metadata)),
                }
                if seq.structure is not None:
                    seg_dict["_structure"] = seq.structure
                if seq.logits is not None:
                    seg_dict["_logits"] = seq.logits
                structured_segments.append(seg_dict)
            structured_constructs.append(
                {
                    "label": construct.label,
                    "type": construct.sequence_type,
                    "segments": structured_segments,
                }
            )
        results.append(
            {
                "result_idx": result_idx,
                "energy_score": make_json_safe(energy_scores[result_idx]),
                "constructs": structured_constructs,
            }
        )

    def get_score(i: int) -> float:
        score: float | None = results[i]["energy_score"]
        return float("inf") if score is None else score

    best_idx = min(range(len(results)), key=get_score) if results else 0
    return {"results": results, "best_result_idx": best_idx}


def build_proposal_results(
    constructs: list[Any],
    outcomes: list[str],
    energy_scores: list[float] | None = None,
) -> list[dict[str, Any]]:
    """Build per-proposal results with accept/reject status from live Construct objects.

    Reads from ``proposal_sequences`` (all proposed sequences) and annotates each
    with whether it was accepted, the rejection reason (if any), and energy score.

    Args:
        constructs (list[Any]): List of Construct objects.
        outcomes (list[str]): Per-proposal outcome, either ``"accepted"`` or a rejection reason string.
        energy_scores (list[float] | None): Per-proposal energy scores. Inf/NaN converted to None.

    Returns:
        list[dict[str, Any]]: List of proposal dicts::

            [
                {
                    "proposal_idx": 0,
                    "accepted": True,
                    "rejected_by": None,
                    "energy_score": 0.42,
                    "constructs": [
                        {
                            "label": "construct_0",
                            "type": "dna",
                            "segments": [
                                {
                                    "label": "promoter",
                                    "sequence": "ATCG",
                                    "constraints": {...},
                                    "generators": {...},
                                    "metadata": {},
                                }
                            ],
                        }
                    ],
                },
                ...,
            ]
    """
    if not constructs or not constructs[0].segments:
        return []

    num_proposals = len(constructs[0].segments[0].proposal_sequences)
    proposal_results = []

    for prop_idx in range(num_proposals):
        structured_constructs = []
        for construct in constructs:
            structured_segments = []
            for seg_idx, segment in enumerate(construct.segments):
                seq = segment.proposal_sequences[prop_idx]
                seg_dict: dict[str, Any] = {
                    "label": segment.label or f"segment_{seg_idx}",
                    "sequence": seq.sequence,
                    "constraints": make_json_safe(copy.deepcopy(seq._constraints_metadata)),
                    "generators": make_json_safe(copy.deepcopy(seq._generator_metadata)),
                    "metadata": make_json_safe(copy.deepcopy(seq._metadata)),
                }
                if seq.structure is not None:
                    seg_dict["_structure"] = seq.structure
                if seq.logits is not None:
                    seg_dict["_logits"] = seq.logits
                structured_segments.append(seg_dict)
            structured_constructs.append(
                {
                    "label": construct.label,
                    "type": construct.sequence_type,
                    "segments": structured_segments,
                }
            )
        if prop_idx >= len(outcomes):
            raise ValueError(
                f"outcomes has {len(outcomes)} entries but there are {num_proposals} proposals; lengths must match"
            )
        if energy_scores is not None and prop_idx >= len(energy_scores):
            raise ValueError(
                f"energy_scores has {len(energy_scores)} entries but there are {num_proposals} proposals; lengths must match"
            )
        outcome = outcomes[prop_idx]
        energy = make_json_safe(energy_scores[prop_idx]) if energy_scores is not None else None
        proposal_results.append(
            {
                "proposal_idx": prop_idx,
                "accepted": outcome == "accepted",
                "rejected_by": None if outcome == "accepted" else outcome,
                "energy_score": energy,
                "constructs": structured_constructs,
            }
        )

    return proposal_results


# =============================================================================
# Shared helpers
# =============================================================================


def _serialize_value(value: Any) -> Any:
    """Coerce complex values to CSV/JSON-friendly scalars.

    Lists/tuples and dicts → JSON string; scalars passthrough.
    """
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value)
    return value


def _collect_all_columns(rows: list[dict[str, Any]]) -> list[str]:
    """Collect all unique column names from rows, preserving insertion order."""
    columns = []
    seen = set()
    for row in rows:
        for key in row:
            if key not in seen:
                columns.append(key)
                seen.add(key)
    return columns


def _flatten_constraint_columns(constraints: dict[str, dict[str, Any]], prefix: str = "") -> dict[str, Any]:
    """Flatten all constraint data with {prefix}{label}.{field} namespacing.

    Used by flatten_sequences, flatten_constructs, flatten_optimization.
    Includes score, weight, weighted_score, all data fields, and multi-segment info.

    Args:
        constraints (dict[str, dict[str, Any]]): Dict mapping constraint labels to their data.
        prefix (str): Column name prefix (e.g., "promoter." for construct-level).
    """
    flat = {}
    for label, cdata in constraints.items():
        base = f"{prefix}{label}"
        for key in ("score", "weight", "weighted_score"):
            if key in cdata:
                flat[f"{base}.{key}"] = cdata[key]
        for key in ("input_segments", "position_in_inputs"):
            if key in cdata:
                flat[f"{base}.{key}"] = _serialize_value(cdata[key])
        for k, v in cdata.get("data", {}).items():
            flat[f"{base}.{k}"] = _serialize_value(v)
    return flat


def _flatten_generator_columns(generators: dict[str, dict[str, Any]], prefix: str = "") -> dict[str, Any]:
    """Flatten generator metadata with {prefix}generator.{key}.{field} namespacing.

    Generator metadata is namespaced by registry key on the Sequence
    (``_generator_metadata[<registry_key>]``); we surface it under a literal
    ``generator.`` prefix so columns can't collide with user-chosen constraint
    labels.

    Args:
        generators (dict[str, dict[str, Any]]): Dict mapping generator registry keys to flat metadata dicts.
        prefix (str): Column name prefix (e.g., "promoter." for construct-level).
    """
    flat = {}
    for key, gdata in generators.items():
        base = f"{prefix}generator.{key}"
        for k, v in gdata.items():
            flat[f"{base}.{k}"] = _serialize_value(v)
    return flat


# =============================================================================
# Flatten functions, one per table
# =============================================================================


def flatten_sequences(
    results: Results,
    segments: set[str] | None = None,
    result_indices: set[int] | None = None,
) -> list[dict[str, Any]]:
    """One row per (result_idx, construct, segment). All constraint fields inline.

    Args:
        results (Results): Output from build_results().
        segments (set[str] | None): If set, only include these segment labels.
        result_indices (set[int] | None): If set, only include these result indices.

    Columns:
        Fixed: result_idx, energy_score, construct, segment, sequence
        Per constraint: {label}.score, {label}.weight, {label}.weighted_score,
            {label}.{data_key}, and optionally {label}.input_segments,
            {label}.position_in_inputs
        Per generator: generator.{registry_key}.{field}
        Metadata: metadata.{key}
    """
    rows = []
    for result_entry in results.get("results", []):
        if result_indices is not None and result_entry["result_idx"] not in result_indices:
            continue
        for construct in result_entry["constructs"]:
            for segment in construct["segments"]:
                if segments is not None and segment["label"] not in segments:
                    continue
                row = {
                    "result_idx": result_entry["result_idx"],
                    "energy_score": result_entry["energy_score"],
                    "construct": construct["label"],
                    "sequence_type": construct["type"],
                    "segment": segment["label"],
                    "sequence": segment["sequence"],
                }
                row.update(_flatten_constraint_columns(segment.get("constraints", {})))
                row.update(_flatten_generator_columns(segment.get("generators", {})))
                for key, value in segment.get("metadata", {}).items():
                    row[f"metadata.{key}"] = _serialize_value(value)
                for path_key in ("structure_path", "logits_path"):
                    if path_key in segment:
                        row[path_key] = segment[path_key]
                rows.append(row)
    return rows


def flatten_constraints(
    results: Results,
    segments: set[str] | None = None,
    constraints: set[str] | None = None,
    result_indices: set[int] | None = None,
) -> list[dict[str, Any]]:
    """One row per (result_idx, construct, segment, constraint). All metrics.

    Args:
        results (Results): Output from build_results().
        segments (set[str] | None): If set, only include these segment labels.
        constraints (set[str] | None): If set, only include these constraint labels.
        result_indices (set[int] | None): If set, only include these result indices.

    Columns:
        Fixed: result_idx, energy_score, construct, segment, constraint
        Standard: score, weight, weighted_score
        Multi-segment (when applicable): input_segments, position_in_inputs
        Custom data: {key} un-prefixed (one constraint per row)
    """
    rows = []
    for result_entry in results.get("results", []):
        if result_indices is not None and result_entry["result_idx"] not in result_indices:
            continue
        for construct in result_entry["constructs"]:
            for segment in construct["segments"]:
                if segments is not None and segment["label"] not in segments:
                    continue
                for label, cdata in segment.get("constraints", {}).items():
                    if constraints is not None and label not in constraints:
                        continue
                    row = {
                        "result_idx": result_entry["result_idx"],
                        "energy_score": result_entry["energy_score"],
                        "construct": construct["label"],
                        "sequence_type": construct["type"],
                        "segment": segment["label"],
                        "constraint": label,
                        "score": cdata.get("score"),
                        "weight": cdata.get("weight"),
                        "weighted_score": cdata.get("weighted_score"),
                    }
                    for key in ("input_segments", "position_in_inputs"):
                        if key in cdata:
                            row[key] = _serialize_value(cdata[key])
                    for k, v in cdata.get("data", {}).items():
                        row[k] = _serialize_value(v)
                    rows.append(row)
    return rows


def flatten_constructs(
    results: Results,
    segments: set[str] | None = None,
    result_indices: set[int] | None = None,
) -> list[dict[str, Any]]:
    """One row per (result_idx, construct). Per-segment data as prefixed columns.

    Args:
        results (Results): Output from build_results().
        segments (set[str] | None): If set, only include these segment labels in per-segment columns.
            full_sequence still reflects all segments for construct integrity.
        result_indices (set[int] | None): If set, only include these result indices.

    Columns:
        Fixed: result_idx, energy_score, construct, full_sequence
        Per segment: {segment}.sequence
        Per segment x constraint: {segment}.{constraint}.score, etc.
        Per segment x generator: {segment}.generator.{registry_key}.{field}
        Per segment metadata: {segment}.metadata.{key}
    """
    rows = []
    for result_entry in results.get("results", []):
        if result_indices is not None and result_entry["result_idx"] not in result_indices:
            continue
        for construct in result_entry["constructs"]:
            row = {
                "result_idx": result_entry["result_idx"],
                "energy_score": result_entry["energy_score"],
                "construct": construct["label"],
                "sequence_type": construct["type"],
                "full_sequence": "".join(s["sequence"] for s in construct["segments"]),
            }
            offset = 0
            for segment in construct["segments"]:
                seg_len = len(segment["sequence"])
                if segments is not None and segment["label"] not in segments:
                    offset += seg_len
                    continue
                seg = segment["label"]
                row[f"{seg}.sequence"] = segment["sequence"]
                row[f"{seg}.start"] = offset
                row[f"{seg}.end"] = offset + seg_len
                row.update(
                    _flatten_constraint_columns(
                        segment.get("constraints", {}),
                        prefix=f"{seg}.",
                    )
                )
                row.update(
                    _flatten_generator_columns(
                        segment.get("generators", {}),
                        prefix=f"{seg}.",
                    )
                )
                for key, value in segment.get("metadata", {}).items():
                    row[f"{seg}.metadata.{key}"] = _serialize_value(value)
                offset += seg_len
            rows.append(row)
    return rows


def flatten_optimization(
    history: list[dict[str, Any]],
    segments: set[str] | None = None,
    result_indices: set[int] | None = None,
    include_proposals: bool = False,
) -> list[dict[str, Any]]:
    """One row per (timepoint, result_idx). Sequences + constraint scores.

    History entries use the same results format as extract_results(),
    so traversal is identical to the other flatten functions.

    Args:
        history (list[dict[str, Any]]): List of history entries from optimizer(s).
        segments (set[str] | None): If set, only include these segment labels.
        result_indices (set[int] | None): If set, only include these result indices.
        include_proposals (bool): If True, add proposal rows alongside result rows.
            Result rows get ``pool="result"``, proposal rows get
            ``pool="proposal"`` with ``proposal_idx``, ``accepted``,
            ``rejected_by`` columns.

    Columns:
        Fixed: timepoint, result_idx, energy_score, optimizer.*
        Single construct: sequence_type, {segment}.sequence, {segment}.{constraint}.score, ...
        Multi-construct: {construct}.sequence_type, {construct}.{segment}.sequence, ...
        When include_proposals=True:
            pool, proposal_idx, accepted, rejected_by, energy_score
    """
    rows = []
    for entry in history:
        timepoint = entry["time_step"]
        optimizer_columns = {f"optimizer.{key}": _serialize_value(value) for key, value in entry["optimizer"].items()}
        for result_entry in entry.get("results", []):
            if result_indices is not None and result_entry["result_idx"] not in result_indices:
                continue
            row = {
                "timepoint": timepoint,
                "result_idx": result_entry["result_idx"],
                "energy_score": result_entry["energy_score"],
            }
            row.update(optimizer_columns)
            if "stage" in entry:
                row["stage"] = entry["stage"]
            if include_proposals:
                row["pool"] = "result"
            multi_construct = len(result_entry["constructs"]) > 1
            for ci, construct in enumerate(result_entry["constructs"]):
                con = construct.get("label") or f"construct_{ci}"
                if multi_construct:
                    row[f"{con}.sequence_type"] = construct.get("type", "")
                else:
                    row["sequence_type"] = construct.get("type", "")
                for segment in construct["segments"]:
                    if segments is not None and segment["label"] not in segments:
                        continue
                    seg = f"{con}.{segment['label']}" if multi_construct else segment["label"]
                    row[f"{seg}.sequence"] = segment["sequence"]
                    row.update(
                        _flatten_constraint_columns(
                            segment.get("constraints", {}),
                            prefix=f"{seg}.",
                        )
                    )
                    row.update(
                        _flatten_generator_columns(
                            segment.get("generators", {}),
                            prefix=f"{seg}.",
                        )
                    )
            rows.append(row)

        # Append proposal rows when requested
        if include_proposals:
            for proposal in entry.get("proposal_results", []):
                row = {
                    "timepoint": timepoint,
                    "pool": "proposal",
                    "proposal_idx": proposal["proposal_idx"],
                    "accepted": proposal["accepted"],
                    "rejected_by": proposal["rejected_by"],
                    "energy_score": proposal.get("energy_score"),
                }
                row.update(optimizer_columns)
                if "stage" in entry:
                    row["stage"] = entry["stage"]
                multi_construct = len(proposal["constructs"]) > 1
                for ci, construct in enumerate(proposal["constructs"]):
                    con = construct.get("label") or f"construct_{ci}"
                    if multi_construct:
                        row[f"{con}.sequence_type"] = construct.get("type", "")
                    else:
                        row["sequence_type"] = construct.get("type", "")
                    for segment in construct["segments"]:
                        if segments is not None and segment["label"] not in segments:
                            continue
                        seg = f"{con}.{segment['label']}" if multi_construct else segment["label"]
                        row[f"{seg}.sequence"] = segment["sequence"]
                        row.update(
                            _flatten_constraint_columns(
                                segment.get("constraints", {}),
                                prefix=f"{seg}.",
                            )
                        )
                        row.update(
                            _flatten_generator_columns(
                                segment.get("generators", {}),
                                prefix=f"{seg}.",
                            )
                        )
                rows.append(row)

    return rows


_ALL_TABLES = ("sequences", "constraints", "constructs", "optimization")


def flatten_table(
    table: str,
    results: Results,
    history: list[dict[str, Any]],
    *,
    segments: set[str] | None = None,
    result_indices: set[int] | None = None,
    constraints: set[str] | None = None,
    include_proposals: bool = False,
) -> list[dict[str, Any]]:
    """Dispatch to the appropriate flatten function for *table*.

    Args:
        table (str): One of ``sequences``, ``constraints``, ``constructs``,
            or ``optimization``.
        results (Results): Output from :func:`build_results`.
        history (list[dict[str, Any]]): Optimization history entries.
        segments (set[str] | None): Only include these segment labels.
        result_indices (set[int] | None): Only include these result indices.
        constraints (set[str] | None): Only include these constraint labels (constraints table only).
        include_proposals (bool): Include proposal rows (optimization table only).

    Raises:
        ValueError: If *table* is not a recognized name.
    """
    filters: dict[str, Any] = {
        "segments": segments,
        "result_indices": result_indices,
    }
    if table == "optimization":
        return flatten_optimization(history, include_proposals=include_proposals, **filters)
    if table == "sequences":
        return flatten_sequences(results, **filters)
    if table == "constraints":
        return flatten_constraints(results, constraints=constraints, **filters)
    if table == "constructs":
        return flatten_constructs(results, **filters)
    raise ValueError(f"Unknown table '{table}'. Choose from: sequences, constraints, constructs, optimization")


# =============================================================================
# Format Writers
# =============================================================================


def to_csv(rows: list[dict[str, Any]], output: Path | IO[str] | None = None) -> str:
    """Write rows to CSV format.

    Args:
        rows (list[dict[str, Any]]): List of dicts with consistent keys
        output (Path | IO[str] | None): Path or file-like object. If None, returns string.

    Returns:
        str: CSV string if output is None
    """
    if not rows:
        return ""

    columns = _collect_all_columns(rows)
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in columns})

    csv_str = buffer.getvalue()

    if output is None:
        return csv_str
    if isinstance(output, Path):
        output.write_text(csv_str)
        return csv_str
    output.write(csv_str)
    return csv_str


def to_tsv(rows: list[dict[str, Any]], output: Path | IO[str] | None = None) -> str:
    """Write rows to TSV format.

    Args:
        rows (list[dict[str, Any]]): List of dicts with consistent keys
        output (Path | IO[str] | None): Path or file-like object. If None, returns string.

    Returns:
        str: TSV string if output is None
    """
    if not rows:
        return ""

    columns = _collect_all_columns(rows)
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, delimiter="\t", extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in columns})

    tsv_str = buffer.getvalue()

    if output is None:
        return tsv_str
    if isinstance(output, Path):
        output.write_text(tsv_str)
        return tsv_str
    output.write(tsv_str)
    return tsv_str


def to_json(
    rows: list[dict[str, Any]],
    output: Path | IO[str] | None = None,
    indent: int = 2,
) -> str:
    """Write rows to JSON format.

    Args:
        rows (list[dict[str, Any]]): List of dicts
        output (Path | IO[str] | None): Path or file-like object. If None, returns string.
        indent (int): JSON indentation (default 2)

    Returns:
        str: JSON string if output is None
    """
    json_str = json.dumps(rows, indent=indent, default=str)

    if output is None:
        return json_str
    if isinstance(output, Path):
        output.write_text(json_str)
        return json_str
    output.write(json_str)
    return json_str


def to_xlsx(rows: list[dict[str, Any]], output: Path | IO[bytes]) -> None:
    """Write rows to Excel format (single sheet).

    Args:
        rows (list[dict[str, Any]]): List of dicts with consistent keys
        output (Path | IO[bytes]): Path or file-like object (required for xlsx)

    """
    from openpyxl import Workbook

    if not rows:
        return

    columns = _collect_all_columns(rows)
    wb = Workbook()
    ws = wb.active

    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    if isinstance(output, Path):
        wb.save(str(output))
    else:
        wb.save(output)


def to_xlsx_workbook(tables: dict[str, list[dict[str, Any]]], output: Path) -> None:
    """Write multiple tables as sheets in a single Excel workbook.

    Args:
        tables (dict[str, list[dict[str, Any]]]): Dict mapping sheet names to row lists.
        output (Path): Output file path.

    """
    from openpyxl import Workbook

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, rows in tables.items():
        ws = wb.create_sheet(title=sheet_name)
        if not rows:
            continue
        columns = _collect_all_columns(rows)
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    wb.save(str(output))


# =============================================================================
# High-level export function
# =============================================================================


def write_export(
    rows: list[dict[str, Any]],
    format: Format,
    path: Path | None = None,
) -> str | None:
    """Write rows to the specified format.

    Args:
        rows (list[dict[str, Any]]): List of dicts to export
        format (Format): Output format ("csv", "tsv", "json", "xlsx")
        path (Path | None): Output path. If None, returns string (not supported for xlsx).

    Returns:
        str | None: String content for csv/tsv/json when path is None, else None
    """
    if format == "csv":
        return to_csv(rows, path)
    if format == "tsv":
        return to_tsv(rows, path)
    if format == "json":
        return to_json(rows, path)
    if format == "xlsx":
        if path is None:
            raise ValueError("xlsx format requires a file path")
        to_xlsx(rows, path)
        return None
    raise ValueError(f"Unsupported format: {format}")


# =============================================================================
# FASTA export
# =============================================================================


def to_fasta(
    results: Results,
    segments: set[str] | None = None,
    result_indices: set[int] | None = None,
    header_format: str = "{construct}_{segment}_result{result_idx}",
    output: Path | IO[str] | None = None,
) -> str:
    """Export sequences in FASTA format for bioinformatics pipelines.

    Args:
        results (Results): Output from build_results().
        segments (set[str] | None): If set, only include these segment labels.
        result_indices (set[int] | None): If set, only include these result indices.
        header_format (str): Python format string for FASTA headers. Available
            fields: construct, segment, result_idx, energy_score, sequence_type.
        output (Path | IO[str] | None): Path or file-like object. If None, returns string.

    Returns:
        str: FASTA string if output is None.
    """
    lines: list[str] = []
    for result_entry in results.get("results", []):
        if result_indices is not None and result_entry["result_idx"] not in result_indices:
            continue
        for construct in result_entry["constructs"]:
            for segment in construct["segments"]:
                if segments is not None and segment["label"] not in segments:
                    continue
                header = header_format.format(
                    construct=construct["label"],
                    segment=segment["label"],
                    result_idx=result_entry["result_idx"],
                    energy_score=result_entry["energy_score"],
                    sequence_type=construct.get("type", ""),
                )
                lines.append(f">{header}")
                lines.append(segment["sequence"])

    fasta_str = "\n".join(lines) + "\n" if lines else ""

    if output is None:
        return fasta_str
    if isinstance(output, Path):
        output.write_text(fasta_str)
        return fasta_str
    output.write(fasta_str)
    return fasta_str


# =============================================================================
# Folder export
# =============================================================================

_ASSETS_DIR_NAME = "assets"


def write_results_folder(
    *,
    results: Results,
    path: Path | str,
    history: list[dict[str, Any]] | None = None,
    format: Format = "csv",
    include_proposals: bool = False,
    segments: set[str] | None = None,
    result_indices: set[int] | None = None,
    constraints: set[str] | None = None,
) -> Path:
    """Write 4 tables + FASTA + ``assets/`` to *path* and return the directory.

    Materializes any in-memory ``_structure`` / ``_logits`` on segments into
    ``assets/`` and stamps ``structure_path`` / ``logits_path`` columns in the
    sequences table. xlsx writes a single ``results.xlsx`` workbook inside the
    folder. Filter kwargs forward to :func:`flatten_table`. See
    :meth:`proto_language.language.core.Program.export` for the folder layout.
    """
    out_dir = Path(path)
    out_dir.mkdir(parents=True, exist_ok=True)
    assets_dir = out_dir / _ASSETS_DIR_NAME
    assets_dir.mkdir(exist_ok=True)

    rewritten: Results = copy.deepcopy(results)
    rewritten_history = copy.deepcopy(history) if history else []

    for result_entry in rewritten.get("results", []):
        r_idx = result_entry["result_idx"]
        for c_idx, construct in enumerate(result_entry["constructs"]):
            for s_idx, segment in enumerate(construct["segments"]):
                _materialize_segment_payloads(segment, r_idx, c_idx, s_idx, assets_dir)

    rows_by_table = {
        name: flatten_table(
            name,
            rewritten,
            rewritten_history,
            segments=segments,
            result_indices=result_indices,
            constraints=constraints,
            include_proposals=include_proposals,
        )
        for name in _ALL_TABLES
    }

    if format == "xlsx":
        to_xlsx_workbook(rows_by_table, out_dir / "results.xlsx")
    else:
        for name, rows in rows_by_table.items():
            table_path = out_dir / f"{name}.{format}"
            write_export(rows, format, table_path)
            if not table_path.exists():
                table_path.write_text("")
    to_fasta(rewritten, segments=segments, result_indices=result_indices, output=out_dir / "sequences.fasta")

    return out_dir


def _materialize_segment_payloads(
    segment: dict[str, Any],
    r_idx: int,
    c_idx: int,
    s_idx: int,
    assets_dir: Path,
) -> None:
    """Pop ``_structure`` / ``_logits`` from *segment*, write them to *assets_dir*, and stamp path columns."""
    structure = segment.pop("_structure", None)
    if structure is not None:
        fmt = getattr(structure, "structure_format", None) or "pdb"
        ext = ".cif" if fmt == "cif" else ".pdb"
        fname = f"res{r_idx}_con{c_idx}_seg{s_idx}_structure{ext}"
        (assets_dir / fname).write_text(structure.structure)
        segment["structure_path"] = f"{_ASSETS_DIR_NAME}/{fname}"

    logits = segment.pop("_logits", None)
    if logits is not None:
        fname = f"res{r_idx}_con{c_idx}_seg{s_idx}_logits.npy"
        np.save(assets_dir / fname, logits)
        segment["logits_path"] = f"{_ASSETS_DIR_NAME}/{fname}"
