"""
Export utilities for optimization results at different granularities.

Five tables, each with a single natural format:
- sequences:    One row per (result_idx, construct, segment)
- constraints:  One row per (result_idx, construct, segment, constraint)
- constructs:   One row per (result_idx, construct)
- optimization: One row per (timepoint, result_idx)
- fasta:        Standard FASTA format for bioinformatics pipelines

Supports CSV, TSV, JSON, FASTA, and Excel output formats.
"""

from __future__ import annotations

import copy
import csv
import json
from io import StringIO
from pathlib import Path
from typing import IO, Any, Callable, Dict, List, Literal, Optional, Set, Union

from proto_language.utils.helpers import filter_inf_nan_scores

# Type aliases
Format = Literal["csv", "tsv", "json", "xlsx"]
Results = Dict[str, Any]  # Output from build_results()


# =============================================================================
# Build results
# =============================================================================


def build_results(
    constructs: list,
    energy_scores: List[float],
) -> Results:
    """Build standardized results from live Construct objects.

    Produces the canonical format consumed by all flatten/export functions.
    Infinite/NaN energy scores are converted to None for JSON compatibility.

    Args:
        constructs: List of Construct objects.
        energy_scores: List of energy scores (one per result).

    Returns:
        Dict with "results" (list of result dicts) and "best_result_idx"::

            {
                "results": [{
                    "result_idx": 0,
                    "energy_score": 0.5,
                    "constructs": [{
                        "label": "construct_0",
                        "type": "dna",
                        "segments": [{
                            "label": "promoter",
                            "sequence": "ATCG",
                            "constraints": {
                                "gc_content": {
                                    "score": 0.5,
                                    "weight": 1.0,
                                    "weighted_score": 0.5,
                                    "data": {"gc_content": 50.0}
                                }
                            },
                            "metadata": {}
                        }]
                    }]
                }],
                "best_result_idx": 0
            }
    """
    if not constructs or not constructs[0].segments:
        return {"results": [], "best_result_idx": 0}

    num_results = len(constructs[0].segments[0].result_sequences)
    results = []

    for result_idx in range(num_results):
        structured_constructs = []
        for construct in constructs:
            structured_segments = []
            for seg_idx, segment in enumerate(construct.segments):
                seq = segment.result_sequences[result_idx]
                structured_segments.append({
                    "label": segment.label or f"segment_{seg_idx}",
                    "sequence": seq.sequence,
                    "constraints": copy.deepcopy(seq._constraints_metadata),
                    "metadata": copy.deepcopy(seq._metadata),
                })
            structured_constructs.append({
                "label": construct.label,
                "type": construct.sequence_type,
                "segments": structured_segments,
            })
        results.append({
            "result_idx": result_idx,
            "energy_score": filter_inf_nan_scores(energy_scores[result_idx]),
            "constructs": structured_constructs,
        })

    def get_score(i: int) -> float:
        score = results[i]["energy_score"]
        return float("inf") if score is None else score

    best_idx = (
        min(range(len(results)), key=get_score) if results else 0
    )
    return {"results": results, "best_result_idx": best_idx}


def build_candidate_results(
    constructs: list,
    outcomes: list[str],
    energy_scores: list[float] | None = None,
) -> list[dict[str, Any]]:
    """Build per-candidate results with accept/reject status from live Construct objects.

    Reads from ``candidate_sequences`` (all proposed sequences) and annotates each
    with whether it was accepted, the rejection reason (if any), and energy score.

    Args:
        constructs: List of Construct objects.
        outcomes: Per-candidate outcome — ``"accepted"`` or a rejection reason string.
        energy_scores: Per-candidate energy scores. Inf/NaN converted to None.

    Returns:
        List of candidate dicts::

            [{
                "candidate_idx": 0,
                "accepted": True,
                "rejected_by": None,
                "energy_score": 0.42,
                "constructs": [{
                    "label": "construct_0",
                    "type": "dna",
                    "segments": [{
                        "label": "promoter",
                        "sequence": "ATCG",
                        "constraints": {...},
                        "metadata": {}
                    }]
                }]
            }, ...]
    """
    if not constructs or not constructs[0].segments:
        return []

    num_candidates = len(constructs[0].segments[0].candidate_sequences)
    candidate_results = []

    for cand_idx in range(num_candidates):
        structured_constructs = []
        for construct in constructs:
            structured_segments = []
            for seg_idx, segment in enumerate(construct.segments):
                seq = segment.candidate_sequences[cand_idx]
                structured_segments.append({
                    "label": segment.label or f"segment_{seg_idx}",
                    "sequence": seq.sequence,
                    "constraints": copy.deepcopy(seq._constraints_metadata),
                    "metadata": copy.deepcopy(seq._metadata),
                })
            structured_constructs.append({
                "label": construct.label,
                "type": construct.sequence_type,
                "segments": structured_segments,
            })
        if cand_idx >= len(outcomes):
            raise ValueError(f"outcomes has {len(outcomes)} entries but there are {num_candidates} candidates — lengths must match")
        if energy_scores is not None and cand_idx >= len(energy_scores):
            raise ValueError(f"energy_scores has {len(energy_scores)} entries but there are {num_candidates} candidates — lengths must match")
        outcome = outcomes[cand_idx]
        energy = (
            filter_inf_nan_scores(energy_scores[cand_idx])
            if energy_scores is not None
            else None
        )
        candidate_results.append({
            "candidate_idx": cand_idx,
            "accepted": outcome == "accepted",
            "rejected_by": None if outcome == "accepted" else outcome,
            "energy_score": energy,
            "constructs": structured_constructs,
        })

    return candidate_results


# =============================================================================
# Shared helpers
# =============================================================================


def _serialize_value(value: Any) -> Any:
    """Coerce complex values to CSV/JSON-friendly scalars.

    - FileReference dicts (``__file_ref__: True``) → URL string
    - Lists/tuples → JSON string
    - Other dicts → JSON string
    - Scalars → passthrough
    """
    if isinstance(value, dict):
        if value.get("__file_ref__"):
            return value.get("url", "")
        return json.dumps(value)
    if isinstance(value, (list, tuple)):
        return json.dumps(value)
    return value


def _collect_all_columns(rows: List[Dict]) -> List[str]:
    """Collect all unique column names from rows, preserving insertion order."""
    columns = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                columns.append(key)
                seen.add(key)
    return columns


def _flatten_constraint_columns(
    constraints: Dict[str, Dict], prefix: str = ""
) -> Dict[str, Any]:
    """Flatten all constraint data with {prefix}{label}.{field} namespacing.

    Used by flatten_sequences, flatten_constructs, flatten_optimization.
    Includes score, weight, weighted_score, all data fields, and multi-segment info.

    Args:
        constraints: Dict mapping constraint labels to their data.
        prefix: Column name prefix (e.g., "promoter." for construct-level).
    """
    flat = {}
    for label, cdata in constraints.items():
        base = f"{prefix}{label}"
        for key in ("score", "weight", "weighted_score"):
            if key in cdata:
                flat[f"{base}.{key}"] = cdata[key]
        for key in ("input_segments", "position_in_inputs"):
            if key in cdata:
                flat[f"{base}.{key}"] = _serialize_value(cdata[key])
        for k, v in cdata.get("data", {}).items():
            flat[f"{base}.{k}"] = _serialize_value(v)
    return flat


# =============================================================================
# Flatten functions — one per table
# =============================================================================


def flatten_sequences(
    results: Results,
    segments: Optional[Set[str]] = None,
    result_indices: Optional[Set[int]] = None,
) -> List[Dict[str, Any]]:
    """One row per (result_idx, construct, segment). All constraint fields inline.

    Args:
        results: Output from build_results().
        segments: If set, only include these segment labels.
        result_indices: If set, only include these result indices.

    Columns:
        Fixed: result_idx, energy_score, construct, segment, sequence
        Per constraint: {label}.score, {label}.weight, {label}.weighted_score,
            {label}.{data_key}, and optionally {label}.input_segments,
            {label}.position_in_inputs
        Metadata: metadata.{key}
    """
    rows = []
    for result_entry in results.get("results", []):
        if result_indices is not None and result_entry["result_idx"] not in result_indices:
            continue
        for construct in result_entry["constructs"]:
            for segment in construct["segments"]:
                if segments is not None and segment["label"] not in segments:
                    continue
                row = {
                    "result_idx": result_entry["result_idx"],
                    "energy_score": result_entry["energy_score"],
                    "construct": construct["label"],
                    "sequence_type": construct["type"],
                    "segment": segment["label"],
                    "sequence": segment["sequence"],
                }
                row.update(
                    _flatten_constraint_columns(
                        segment.get("constraints", {})
                    )
                )
                for key, value in segment.get("metadata", {}).items():
                    row[f"metadata.{key}"] = _serialize_value(value)
                rows.append(row)
    return rows


def flatten_constraints(
    results: Results,
    segments: Optional[Set[str]] = None,
    constraints: Optional[Set[str]] = None,
    result_indices: Optional[Set[int]] = None,
) -> List[Dict[str, Any]]:
    """One row per (result_idx, construct, segment, constraint). All metrics.

    Args:
        results: Output from build_results().
        segments: If set, only include these segment labels.
        constraints: If set, only include these constraint labels.
        result_indices: If set, only include these result indices.

    Columns:
        Fixed: result_idx, energy_score, construct, segment, constraint
        Standard: score, weight, weighted_score
        Multi-segment (when applicable): input_segments, position_in_inputs
        Custom data: {key} un-prefixed (one constraint per row)
    """
    rows = []
    for result_entry in results.get("results", []):
        if result_indices is not None and result_entry["result_idx"] not in result_indices:
            continue
        for construct in result_entry["constructs"]:
            for segment in construct["segments"]:
                if segments is not None and segment["label"] not in segments:
                    continue
                for label, cdata in segment.get("constraints", {}).items():
                    if constraints is not None and label not in constraints:
                        continue
                    row = {
                        "result_idx": result_entry["result_idx"],
                        "energy_score": result_entry["energy_score"],
                        "construct": construct["label"],
                        "sequence_type": construct["type"],
                        "segment": segment["label"],
                        "constraint": label,
                        "score": cdata.get("score"),
                        "weight": cdata.get("weight"),
                        "weighted_score": cdata.get("weighted_score"),
                    }
                    for key in ("input_segments", "position_in_inputs"):
                        if key in cdata:
                            row[key] = _serialize_value(cdata[key])
                    for k, v in cdata.get("data", {}).items():
                        row[k] = _serialize_value(v)
                    rows.append(row)
    return rows


def flatten_constructs(
    results: Results,
    segments: Optional[Set[str]] = None,
    result_indices: Optional[Set[int]] = None,
) -> List[Dict[str, Any]]:
    """One row per (result_idx, construct). Per-segment data as prefixed columns.

    Args:
        results: Output from build_results().
        segments: If set, only include these segment labels in per-segment columns.
            full_sequence still reflects all segments for construct integrity.
        result_indices: If set, only include these result indices.

    Columns:
        Fixed: result_idx, energy_score, construct, full_sequence
        Per segment: {segment}.sequence
        Per segment x constraint: {segment}.{constraint}.score, etc.
        Per segment metadata: {segment}.metadata.{key}
    """
    rows = []
    for result_entry in results.get("results", []):
        if result_indices is not None and result_entry["result_idx"] not in result_indices:
            continue
        for construct in result_entry["constructs"]:
            row = {
                "result_idx": result_entry["result_idx"],
                "energy_score": result_entry["energy_score"],
                "construct": construct["label"],
                "sequence_type": construct["type"],
                "full_sequence": "".join(
                    s["sequence"] for s in construct["segments"]
                ),
            }
            offset = 0
            for segment in construct["segments"]:
                seg_len = len(segment["sequence"])
                if segments is not None and segment["label"] not in segments:
                    offset += seg_len
                    continue
                seg = segment["label"]
                row[f"{seg}.sequence"] = segment["sequence"]
                row[f"{seg}.start"] = offset
                row[f"{seg}.end"] = offset + seg_len
                row.update(
                    _flatten_constraint_columns(
                        segment.get("constraints", {}),
                        prefix=f"{seg}.",
                    )
                )
                for key, value in segment.get("metadata", {}).items():
                    row[f"{seg}.metadata.{key}"] = _serialize_value(value)
                offset += seg_len
            rows.append(row)
    return rows


def flatten_optimization(
    history: List[Dict[str, Any]],
    segments: Optional[Set[str]] = None,
    result_indices: Optional[Set[int]] = None,
    include_candidates: bool = False,
) -> List[Dict[str, Any]]:
    """One row per (timepoint, result_idx). Sequences + constraint scores.

    History entries use the same results format as extract_results(),
    so traversal is identical to the other flatten functions.

    Args:
        history: List of history entries from optimizer(s).
        segments: If set, only include these segment labels.
        result_indices: If set, only include these result indices.
        include_candidates: If True, add candidate rows alongside result rows.
            Result rows get ``pool="result"``, candidate rows get
            ``pool="candidate"`` with ``candidate_idx``, ``accepted``,
            ``rejected_by`` columns. When False, output is identical to
            previous behavior (no new columns).

    Columns:
        Fixed: timepoint, result_idx, energy_score
        Single construct: sequence_type, {segment}.sequence, {segment}.{constraint}.score, ...
        Multi-construct: {construct}.sequence_type, {construct}.{segment}.sequence, ...
        When include_candidates=True:
            pool, candidate_idx, accepted, rejected_by, energy_score
    """
    rows = []
    for entry in history:
        timepoint = entry["time_step"]
        for result_entry in entry.get("results", []):
            if result_indices is not None and result_entry["result_idx"] not in result_indices:
                continue
            row = {
                "timepoint": timepoint,
                "result_idx": result_entry["result_idx"],
                "energy_score": result_entry["energy_score"],
            }
            if "stage" in entry:
                row["stage"] = entry["stage"]
            if include_candidates:
                row["pool"] = "result"
            multi_construct = len(result_entry["constructs"]) > 1
            for ci, construct in enumerate(result_entry["constructs"]):
                con = construct.get("label") or f"construct_{ci}"
                if multi_construct:
                    row[f"{con}.sequence_type"] = construct.get("type", "")
                else:
                    row["sequence_type"] = construct.get("type", "")
                for segment in construct["segments"]:
                    if segments is not None and segment["label"] not in segments:
                        continue
                    seg = f"{con}.{segment['label']}" if multi_construct else segment["label"]
                    row[f"{seg}.sequence"] = segment["sequence"]
                    row.update(
                        _flatten_constraint_columns(
                            segment.get("constraints", {}),
                            prefix=f"{seg}.",
                        )
                    )
            rows.append(row)

        # Append candidate rows when requested
        if include_candidates:
            for candidate in entry.get("candidate_results", []):
                row = {
                    "timepoint": timepoint,
                    "pool": "candidate",
                    "candidate_idx": candidate["candidate_idx"],
                    "accepted": candidate["accepted"],
                    "rejected_by": candidate["rejected_by"],
                    "energy_score": candidate.get("energy_score"),
                }
                if "stage" in entry:
                    row["stage"] = entry["stage"]
                multi_construct = len(candidate["constructs"]) > 1
                for ci, construct in enumerate(candidate["constructs"]):
                    con = construct.get("label") or f"construct_{ci}"
                    if multi_construct:
                        row[f"{con}.sequence_type"] = construct.get("type", "")
                    else:
                        row["sequence_type"] = construct.get("type", "")
                    for segment in construct["segments"]:
                        if segments is not None and segment["label"] not in segments:
                            continue
                        seg = f"{con}.{segment['label']}" if multi_construct else segment["label"]
                        row[f"{seg}.sequence"] = segment["sequence"]
                        row.update(
                            _flatten_constraint_columns(
                                segment.get("constraints", {}),
                                prefix=f"{seg}.",
                            )
                        )
                rows.append(row)

    return rows


_ALL_TABLES = ("sequences", "constraints", "constructs", "optimization")


def flatten_table(
    table: str,
    results: Results,
    history: List[Dict[str, Any]],
    *,
    segments: Optional[Set[str]] = None,
    result_indices: Optional[Set[int]] = None,
    constraints: Optional[Set[str]] = None,
    include_candidates: bool = False,
) -> List[Dict[str, Any]]:
    """Dispatch to the appropriate flatten function for *table*.

    Args:
        table: One of ``sequences``, ``constraints``, ``constructs``,
            or ``optimization``.
        results: Output from :func:`build_results`.
        history: Optimization history entries.
        segments: Only include these segment labels.
        result_indices: Only include these result indices.
        constraints: Only include these constraint labels (constraints table only).
        include_candidates: Include candidate rows (optimization table only).

    Raises:
        ValueError: If *table* is not a recognized name.
    """
    filters = {"segments": segments, "result_indices": result_indices}
    if table == "optimization":
        return flatten_optimization(
            history, include_candidates=include_candidates, **filters
        )
    if table == "sequences":
        return flatten_sequences(results, **filters)
    if table == "constraints":
        return flatten_constraints(
            results, constraints=constraints, **filters
        )
    if table == "constructs":
        return flatten_constructs(results, **filters)
    raise ValueError(
        f"Unknown table '{table}'. "
        f"Choose from: sequences, constraints, constructs, optimization"
    )


# =============================================================================
# Format Writers
# =============================================================================


def to_csv(rows: List[Dict], output: Union[Path, IO, None] = None) -> str:
    """Write rows to CSV format.

    Args:
        rows: List of dicts with consistent keys
        output: Path or file-like object. If None, returns string.

    Returns:
        CSV string if output is None
    """
    if not rows:
        return ""

    columns = _collect_all_columns(rows)
    buffer = StringIO()
    writer = csv.DictWriter(
        buffer, fieldnames=columns, extrasaction="ignore"
    )
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in columns})

    csv_str = buffer.getvalue()

    if output is None:
        return csv_str
    elif isinstance(output, Path):
        output.write_text(csv_str)
        return csv_str
    else:
        output.write(csv_str)
        return csv_str


def to_tsv(rows: List[Dict], output: Union[Path, IO, None] = None) -> str:
    """Write rows to TSV format.

    Args:
        rows: List of dicts with consistent keys
        output: Path or file-like object. If None, returns string.

    Returns:
        TSV string if output is None
    """
    if not rows:
        return ""

    columns = _collect_all_columns(rows)
    buffer = StringIO()
    writer = csv.DictWriter(
        buffer, fieldnames=columns, delimiter="\t", extrasaction="ignore"
    )
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in columns})

    tsv_str = buffer.getvalue()

    if output is None:
        return tsv_str
    elif isinstance(output, Path):
        output.write_text(tsv_str)
        return tsv_str
    else:
        output.write(tsv_str)
        return tsv_str


def to_json(
    rows: List[Dict],
    output: Union[Path, IO, None] = None,
    indent: int = 2,
) -> str:
    """Write rows to JSON format.

    Args:
        rows: List of dicts
        output: Path or file-like object. If None, returns string.
        indent: JSON indentation (default 2)

    Returns:
        JSON string if output is None
    """
    json_str = json.dumps(rows, indent=indent, default=str)

    if output is None:
        return json_str
    elif isinstance(output, Path):
        output.write_text(json_str)
        return json_str
    else:
        output.write(json_str)
        return json_str


def to_xlsx(rows: List[Dict], output: Union[Path, IO]) -> None:
    """Write rows to Excel format (single sheet).

    Args:
        rows: List of dicts with consistent keys
        output: Path or file-like object (required for xlsx)

    """
    from openpyxl import Workbook

    if not rows:
        return

    columns = _collect_all_columns(rows)
    wb = Workbook()
    ws = wb.active

    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    if isinstance(output, Path):
        wb.save(str(output))
    else:
        wb.save(output)


def to_xlsx_workbook(
    tables: Dict[str, List[Dict]], output: Path
) -> None:
    """Write multiple tables as sheets in a single Excel workbook.

    Args:
        tables: Dict mapping sheet names to row lists.
        output: Output file path.

    """
    from openpyxl import Workbook

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, rows in tables.items():
        ws = wb.create_sheet(title=sheet_name)
        if not rows:
            continue
        columns = _collect_all_columns(rows)
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(
                    row=row_idx, column=col_idx, value=row.get(col_name, "")
                )

    wb.save(str(output))


# =============================================================================
# High-level export function
# =============================================================================


def write_export(
    rows: List[Dict],
    format: Format,
    path: Optional[Path] = None,
) -> Union[str, None]:
    """Write rows to the specified format.

    Args:
        rows: List of dicts to export
        format: Output format ("csv", "tsv", "json", "xlsx")
        path: Output path. If None, returns string (not supported for xlsx).

    Returns:
        String content for csv/tsv/json when path is None, else None
    """
    if format == "csv":
        return to_csv(rows, path)
    elif format == "tsv":
        return to_tsv(rows, path)
    elif format == "json":
        return to_json(rows, path)
    elif format == "xlsx":
        if path is None:
            raise ValueError("xlsx format requires a file path")
        to_xlsx(rows, path)
        return None
    else:
        raise ValueError(f"Unsupported format: {format}")


def export_tables(
    flatten_fn: Callable[[str], List[Dict[str, Any]]],
    path: Path | str,
    format: Format,
    table: str | None = None,
) -> Path:
    """Write one or all result tables to *path*.

    Without *table*: writes all 4 tables. csv/tsv/json produce a directory
    with one file per table; xlsx produces a single workbook with 4 sheets.
    With *table*: writes a single file to *path*.

    Args:
        flatten_fn: Called with a table name, returns flattened rows.
        path: Output directory (all tables) or file (single table / xlsx).
        format: Output format.
        table: Single table name, or ``None`` to export all.
    """
    path = Path(path)
    if table is not None:
        write_export(flatten_fn(table), format, path)
        return path
    all_tables = {name: flatten_fn(name) for name in _ALL_TABLES}
    if format == "xlsx":
        to_xlsx_workbook(all_tables, path)
    else:
        path.mkdir(parents=True, exist_ok=True)
        for name, rows in all_tables.items():
            write_export(rows, format, path / f"{name}.{format}")
    return path


# =============================================================================
# FASTA export
# =============================================================================


def to_fasta(
    results: Results,
    segments: Optional[Set[str]] = None,
    result_indices: Optional[Set[int]] = None,
    header_format: str = "{construct}_{segment}_result{result_idx}",
    output: Union[Path, IO, None] = None,
) -> str:
    """Export sequences in FASTA format for bioinformatics pipelines.

    Args:
        results: Output from build_results().
        segments: If set, only include these segment labels.
        result_indices: If set, only include these result indices.
        header_format: Python format string for FASTA headers. Available
            fields: construct, segment, result_idx, energy_score, sequence_type.
        output: Path or file-like object. If None, returns string.

    Returns:
        FASTA string if output is None.
    """
    lines: List[str] = []
    for result_entry in results.get("results", []):
        if result_indices is not None and result_entry["result_idx"] not in result_indices:
            continue
        for construct in result_entry["constructs"]:
            for segment in construct["segments"]:
                if segments is not None and segment["label"] not in segments:
                    continue
                header = header_format.format(
                    construct=construct["label"],
                    segment=segment["label"],
                    result_idx=result_entry["result_idx"],
                    energy_score=result_entry["energy_score"],
                    sequence_type=construct.get("type", ""),
                )
                lines.append(f">{header}")
                lines.append(segment["sequence"])

    fasta_str = "\n".join(lines) + "\n" if lines else ""

    if output is None:
        return fasta_str
    elif isinstance(output, Path):
        output.write_text(fasta_str)
        return fasta_str
    else:
        output.write(fasta_str)
        return fasta_str
